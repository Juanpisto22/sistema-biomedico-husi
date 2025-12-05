"""
Funciones de exportación de datos (PDF y Excel).

Este módulo maneja la generación de reportes y exportación
de datos del historial de rondas en diferentes formatos.
"""
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime
import os

try:
    from xhtml2pdf import pisa
    XHTML2PDF_AVAILABLE = True
except ImportError:
    pisa = None
    XHTML2PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from .models import RoundEntry, DailySurgeryRecord


@login_required
@permission_required('rondas.delete_roundentry', raise_exception=True)
def exportar_historial_pdf(request):
    """
    Exporta el historial de rondas a formato PDF.
    
    Esta función genera un documento PDF con todos los registros de rondas,
    incluyendo información detallada de cada servicio y las firmas asociadas.
    
    Requiere permisos de administrador (delete_roundentry).
    
    Returns:
        HttpResponse: Archivo PDF para descarga o mensaje de error
    """
    # Verificar permisos de administrador
    if not request.user.groups.filter(name='Administradores').exists():
        messages.error(request, 'No tienes permisos para exportar datos. Solo los administradores pueden realizar esta acción.')
        return redirect('historial_servicios')
    
    # Verificar disponibilidad de librería PDF
    if not XHTML2PDF_AVAILABLE:
        messages.error(request, 'La funcionalidad de exportación PDF no está disponible. Por favor contacta al administrador del sistema.')
        return redirect('historial_servicios')
    
    try:
        # Obtener todos los registros
        registros_servicios = RoundEntry.objects.select_related("usuario").order_by('-fecha_creacion')
        registros_cirugias = DailySurgeryRecord.objects.select_related("usuario").order_by('-fecha_creacion')
        
        from itertools import chain
        registros_combinados = list(chain(registros_servicios, registros_cirugias))
        
        # Preparar contexto para el template
        context = {
            'registros': registros_combinados,
            'fecha_generacion': timezone.now(),
            'usuario': request.user
        }
        
        # Renderizar HTML desde template
        html_string = render_to_string('rondas/pdf_historial.html', context)
        
        # Crear respuesta HTTP con tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="historial_rondas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Generar PDF desde HTML
        pisa_status = pisa.CreatePDF(html_string, dest=response)
        
        # Verificar errores en la generación
        if pisa_status.err:
            messages.error(request, 'Error al generar el PDF. Por favor intenta nuevamente.')
            return redirect('historial_servicios')
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar a PDF: {str(e)}')
        return redirect('historial_servicios')


@login_required
@permission_required('rondas.delete_roundentry', raise_exception=True)
def exportar_historial_excel(request):
    """
    Exporta el historial de rondas a formato Excel.
    
    Genera un archivo Excel (.xlsx) con todos los registros organizados
    en una tabla con formato profesional, incluyendo encabezados con estilo
    y columnas ajustadas automáticamente.
    
    Requiere permisos de administrador (delete_roundentry).
    
    Returns:
        HttpResponse: Archivo Excel para descarga o mensaje de error
    """
    # Verificar permisos de administrador
    if not request.user.groups.filter(name='Administradores').exists():
        messages.error(request, 'No tienes permisos para exportar datos. Solo los administradores pueden realizar esta acción.')
        return redirect('historial_servicios')
    
    # Verificar disponibilidad de librería Excel
    if not openpyxl:
        messages.error(request, 'La funcionalidad de exportación Excel no está disponible. Por favor contacta al administrador del sistema.')
        return redirect('historial_servicios')
    
    try:
        # Crear nuevo libro de Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Historial Rondas"
        
        # Definir encabezados
        headers = [
            'Fecha', 'Hora', 'Usuario', 'Categoría', 'Servicio',
            'Estado', 'Novedades', 'Observaciones', 'Tipo'
        ]
        
        # Escribir encabezados con formato
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener todos los registros
        registros_servicios = RoundEntry.objects.select_related("usuario").order_by('-fecha_creacion')
        registros_cirugias = DailySurgeryRecord.objects.select_related("usuario").order_by('-fecha_creacion')
        
        # Escribir datos de servicios
        row_num = 2
        for registro in registros_servicios:
            worksheet.cell(row=row_num, column=1).value = registro.fecha_creacion.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=2).value = registro.fecha_creacion.strftime('%H:%M:%S')
            worksheet.cell(row=row_num, column=3).value = registro.usuario.username
            worksheet.cell(row=row_num, column=4).value = registro.get_categoria_display()
            worksheet.cell(row=row_num, column=5).value = registro.subservicio
            worksheet.cell(row=row_num, column=6).value = 'Con eventos' if registro.tiene_eventos_seguridad else 'Sin eventos'
            worksheet.cell(row=row_num, column=7).value = 'Sí' if registro.fuera_de_servicio else 'No'
            worksheet.cell(row=row_num, column=8).value = registro.hallazgo
            worksheet.cell(row=row_num, column=9).value = 'Servicio Regular'
            row_num += 1
        
        # Escribir datos de cirugías
        for registro in registros_cirugias:
            worksheet.cell(row=row_num, column=1).value = registro.fecha.strftime('%Y-%m-%d') if registro.fecha else registro.fecha_creacion.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=2).value = registro.fecha_creacion.strftime('%H:%M:%S')
            worksheet.cell(row=row_num, column=3).value = registro.usuario.username
            worksheet.cell(row=row_num, column=4).value = 'Cirugía'
            worksheet.cell(row=row_num, column=5).value = f"Sala {registro.sala} - {registro.equipo}"
            worksheet.cell(row=row_num, column=6).value = registro.estado_equipo if registro.estado_equipo else 'No especificado'
            worksheet.cell(row=row_num, column=7).value = 'Sí' if registro.equipo_en_uso else 'No'
            worksheet.cell(row=row_num, column=8).value = registro.observaciones
            worksheet.cell(row=row_num, column=9).value = 'Cirugía'
            row_num += 1
        
        # Ajustar ancho de columnas automáticamente
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="historial_rondas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Guardar el archivo Excel en la respuesta
        workbook.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar a Excel: {str(e)}')
        return redirect('historial_servicios')
